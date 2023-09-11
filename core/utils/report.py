import os
import re
import csv
import sys
import traceback
import pandas as pd
import openpyxl as xl
from datetime import date

sys.path.append("/core/utils/")

from mdr import MDR
from log import Logger
from core.utils.helper_helion import Helper
from datetime import timedelta
from docxtpl import DocxTemplate

RAW_CSV = os.getenv("RAW_CSV")
RAW_XLSX = os.getenv("RAW_XLSX")
DOC_REPORT = os.getenv("DOC_REPORT")
DOC_TEMPLATE = os.getenv("DOC_TEMPLATE")
EXCEL_REPORT = os.getenv("EXCEL_REPORT")
YESTERDAY = date.today() - timedelta(days=1)
EXCEL_TEMPLATE = os.getenv("EXCEL_TEMPLATE")

class Report():
    
    def __init__(self, sensors: list, logger: Logger):
        self.logger = logger
        self.content = [] 
        self.sensor_ip = []
        self.count = 0
        self.sensor_list = sensors
        self.today_doc = f"{DOC_REPORT}BGT_Email_Summary_{str(date.today().strftime('%d_%m_%Y'))}.docx"
        self.yesterday_doc = f"{DOC_REPORT}BGT_Email_Summary_{str(YESTERDAY.strftime('%d_%m_%Y'))}.docx"
        self.today_excel = f'{EXCEL_REPORT}BGT-System_Checklist-{str(date.today().strftime("%d/%m/%Y")).replace("/","_")}.xlsx'
        self.yesterday_excel = f'{EXCEL_REPORT}BGT-System_Checklist-{str(YESTERDAY.strftime("%d/%m/%Y")).replace("/","_")}.xlsx'
        
    def export_sensor_list_from_excel(self) -> bool:        
        try: 
            workbook = xl.load_workbook(RAW_XLSX)
            sheet_1 = workbook['Sheet1']
            for i in range(2, sheet_1.max_row + 1):
                self.content.append({
                    'Stt': i - 1,
                    'Sensor': sheet_1.cell(i, 2).value,
                    'IP': sheet_1.cell(i, 3).value,
                    'Location': sheet_1.cell(i, 4).value,
                    'Last_Activity': sheet_1.cell(i, 5).value,
                    'OS_Version':sheet_1.cell(i, 6).value,
                })
            return True
        except Exception as error:
            self.logger.log_message(error, "error")
            traceback.print_exc()
            return False
                
    def generate_offline_sensor_list(self) -> bool:
        if not self.sensor_list:
            return False
        try:
            data = []
            
            for sensor in self.sensor_list:
                if sensor.status == "Online":
                    continue
                
                valid_ip = self.validate_ipv4(sensor.network_adapters)
                valid_zone = self.validate_zone(sensor.group_id)
                
                row = [
                    sensor.computer_name,
                    valid_ip,
                    valid_zone,
                    sensor.last_update,
                    sensor.os_environment_display_string
                ]
                
                data.append(row)
                
                self.sensor_ip.append(valid_ip)
                
            self.count = len(data)  
            
            if not data:
                return False
            
            with open(RAW_CSV, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Sensor', 'IP', 'Location', 'Last Activity', 'OS Version'])
                writer.writerows(data)
                
            read_file = pd.read_csv(RAW_CSV)
            read_file.to_excel(RAW_XLSX, index=True, header=True)
            
            return True
        except Exception as error:
            self.logger.log_message(error, "error")
            traceback.print_exc()
            return False

    def export_docx_report(self):
        try:
            self.logger.log_message("Start export summary doc...", "info")
            doc = DocxTemplate(DOC_TEMPLATE)
            context = {
                'DATE': str(date.today().strftime("%d/%m/%Y")),
                'total_sensor_offline': self.count,
                'table_contents': self.content,
            }
            doc.render(context)
            doc.save(self.today_doc)
        except Exception as error:
            self.logger.log_message(error, "error")
            traceback.print_exc()

    def export_excel_report(self) -> bool:
        try:
            
            total_sensor_offline_str = '\n'.join(self.sensor_ip)
            alert_no_log = Helper(self.logger).get_alert_no_log()
            
            if self.count > 0:
                if len(alert_no_log) > 0:
                    replacement_pair = {"{{LIST_SENSORS_CBR_OFFFLINE}}": "Các địa chỉ IP có sensors offline và còn lại đều online:\n" + str(total_sensor_offline_str),"{{TICK_CBR_OK}}":"","{{TICK_CBR_NOT_OK}}":"x","{{TICK_MGUARD_NO_LOG_OK}}":"","{{TICK_MGUARD_NO_LOG_NOT_OK}}":"x","{{NOTE_ALERT_NO_LOG}}":"Have alert no log"}
                else:
                    replacement_pair = {"{{LIST_SENSORS_CBR_OFFFLINE}}": "Các địa chỉ IP có sensors offline và còn lại đều online:\n" + str(total_sensor_offline_str),"{{TICK_CBR_OK}}":"","{{TICK_CBR_NOT_OK}}":"x","{{TICK_MGUARD_NO_LOG_OK}}":"x","{{TICK_MGUARD_NO_LOG_NOT_OK}}":"","{{NOTE_ALERT_NO_LOG}}":"No alert no log"}
            else:
                if len(alert_no_log) > 0:
                    replacement_pair = {"{{LIST_SENSORS_CBR_OFFFLINE}}": "Các địa chỉ IP có sensors offline và còn lại đều online:\n" + str(total_sensor_offline_str),"{{TICK_CBR_OK}}":"x","{{TICK_CBR_NOT_OK}}":"","{{TICK_MGUARD_NO_LOG_OK}}":"","{{TICK_MGUARD_NO_LOG_NOT_OK}}":"x","{{NOTE_ALERT_NO_LOG}}":"Have alert no log"}
                else:
                    replacement_pair = {"{{LIST_SENSORS_CBR_OFFFLINE}}": "Các địa chỉ IP có sensors offline và còn lại đều online:\n" + str(total_sensor_offline_str),"{{TICK_CBR_OK}}":"x","{{TICK_CBR_NOT_OK}}":"","{{TICK_MGUARD_NO_LOG_OK}}":"x","{{TICK_MGUARD_NO_LOG_NOT_OK}}":"","{{NOTE_ALERT_NO_LOG}}":"No alert no log"}
            
            wb = xl.load_workbook(EXCEL_TEMPLATE)
            
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value in replacement_pair.keys():
                            cell.value = replacement_pair.get(cell.value)
            
            wb.save(self.today_excel)
            return True
        
        except Exception as error:
            self.logger.log_message(error, "error")
            traceback.print_exc()
            return False

    def clear_old_doc(self):
        self.remove_file(self.yesterday_doc)
        self.remove_file(self.yesterday_excel)
    
    def validate_zone(self, group_id: int) -> str:
        return ZONE_MAPPING.get(group_id, 'UNKNOWN')

    def validate_ipv4(self, raw_text: str) -> str:
        ip = IPV4_REGEX.search(raw_text)
        return ip.group(0) if ip else "127.0.0.1"

    def remove_file(self, file_path: str) -> bool:
        if os.path.isfile(file_path): 
            try:
                os.remove(file_path)
                return not os.path.isfile(file_path)   # Return True if file doesn't exist after deleting it.
            except PermissionError:  # Handle permission errors when trying to remove a file that is currently open/locked.
                print(f"Error: Unable to delete {file_path}. File may be in use.")
                return False
        else:
            self.logger.log_message(f"{file_path} file not found", "warning")
            return False   
        
ZONE_MAPPING = {
    2: 'PDC',
    3: 'DRC',
    5: 'DRC',
    6: 'PDC',
    7: 'PDC'
}

IPV4_REGEX = re.compile('(10(\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)){3}|(172\.(1[6-9]|2[0-9]|3[01]))(\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)){2})')

  
if __name__ == "__main__":
    mdr = MDR(Logger())
    sensor_list = mdr.get_sensor_list()
    report = Report(sensors=sensor_list)
    
    report.clear_old_doc()
    
    if report.generate_offline_sensor_list():
        if report.export_sensor_list_from_excel():
            report.export_docx_report()
            report.export_excel_report()
        else:
            print("export from csc error")
    else:
        print("get sensor error")
    
    